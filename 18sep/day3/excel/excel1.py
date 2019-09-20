#!/usr/bin/env python
# coding: utf-8

# In[1]:


#pandas


# In[2]:


from openpyxl import load_workbook


# In[3]:


wba = load_workbook('beforelunch.xlsx')


# In[4]:


type(wba)


# In[5]:


print(wba.sheetnames)


# In[8]:


sheeta = wba['first']


# In[7]:


sheeta = wba.get_sheet_by_name('first')


# In[9]:


print(sheeta['B3'].value)


# In[10]:


var = sheeta['B3']


# In[11]:


for i in range(1, 6):
    print(sheeta.cell(row=i, column=1).value)


# In[13]:


import pandas as pd


# In[16]:


x1 = pd.ExcelFile('beforelunch.xlsx')


# In[15]:


type(dfa)


# In[17]:


print(x1.sheet_names)


# In[18]:


dfa = x1.parse('second')


# In[19]:


dfa


# In[20]:


dicta = {'Data':[6, 7, 8, 90, 12]}


# In[21]:


dfb = pd.DataFrame(dicta)


# In[22]:


dfb


# In[23]:


writer = pd.ExcelWriter('awake.xlsx', engine='xlsxwriter')


# In[24]:


dfb.to_excel(writer, sheet_name='one')


# In[25]:


writer.save()


# In[28]:


from openpyxl import Workbook
wba = Workbook()
wsa = wba.active
wsa['A3'] = 22
wsa['B3'] = 24
wsa['C3'] = "=SUM(A3, B3)"
wba.save('formula.xlsx')


# In[29]:


from openpyxl import Workbook
from openpyxl.drawing.image import Image

wbb = Workbook()
wsb = wbb.active
img = Image('one.jpg')

wsb.add_image(img, 'B2')
wbb.save('image.xlsx')


# In[32]:


from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference

wbb = Workbook()
wsb = wbb.active
for i in range(10):
    wsb.append([i])

values = Reference(wsb, min_col=1, min_row=1, max_col=1, max_row=10)
chart1 = BarChart()

chart1.add_data(values)

wsb.add_chart(chart1, 'C3')
wbb.save('chart.xlsx')


# In[ ]:




